import logging
from typing import Iterator

from datariot.__spi__.error import DataRiotImportException, DataRiotException
from datariot.__spi__.type import Parser, FileFilter, Parsed
from datariot.__util__.io_util import get_files
from datariot.parser.xlsx.__spi__ import XlsxParserConfig, XlsxParsed
from datariot.parser.xlsx.xlsx_model import XlsxRowBox


class XlsxParser(Parser):

    def __init__(self):
        try:
            import openpyxl
        except ImportError:
            raise DataRiotImportException("xlsx")

    def parse(self, path: str, config: XlsxParserConfig = XlsxParserConfig()):
        import openpyxl
        workbook = openpyxl.load_workbook(path)

        if config.worksheet_id is not None:
            dataframe = workbook.worksheets[config.worksheet_id]
        else:
            dataframe = workbook.active
        values = []

        for row in range(0, dataframe.max_row):
            if config.included_rows is not None:
                if (row + 1) not in config.included_rows:
                    continue

            values.append([])
            for col in dataframe.iter_cols(1, dataframe.max_column):
                row_col = col[row]

                if config.included_cols is not None:
                    if row_col.column not in config.included_cols:
                        continue

                values[-1].append(row_col)

        bboxes = [XlsxRowBox(e) for e in values]
        return XlsxParsed(path, bboxes)

    @staticmethod
    def parse_folder(path: str, file_filter: FileFilter = lambda _: True) -> Iterator[Parsed]:
        parser = XlsxParser()
        for file in get_files(path, ".xlsx"):
            try:
                if file_filter(file):
                    yield parser.parse(file)
            except DataRiotException as ex:
                logging.warning(ex)
